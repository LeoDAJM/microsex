import csv
import io
import os
import re
import string
import sys

import FUN.CONF.config_custom as config2
from FUN.CONF.configCC import lang_init
from FUN.CC.portA import *
from FUN.CC.display import *
import rc_icons
from FUN.CC.Editor_Codigo import *
from FUN.CC.Editor_Registros import *
from FUN.CC.Ensamblador import *
from FUN.CC.lst_table import *
from FUN.CC.segments_editor import *
from FUN.CC.Unidad_Control import *
from FUN.CONF.nemonicos import argumentos_instrucciones
from openpyxl import Workbook, load_workbook
from PyQt6.QtCore import Qt, QProcess
from PyQt6.QtGui import QAction, QFont, QFontDatabase, QFontMetricsF, QIcon
from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QMainWindow,
    QMessageBox,
    QSizePolicy,
    QTextEdit,
    QToolBar,
    QVBoxLayout,
    QWidget,
    QTabWidget,
    QGridLayout,
    QSplitter,
)
from FUN.CONF.dict_eng_esp import CC_dict

numeros = tuple(str(i) for i in string.digits)
letras = tuple(str(i) for i in string.ascii_letters)
letras_numeros = letras + numeros
_dict = CC_dict()

class ComputadorCompleto(QMainWindow):
    @property
    def lang_sel(self):
        return self._lang_sel

    @lang_sel.setter
    def lang_sel(self, value):
        self._lang_sel = value
        self._dict_sel = _dict.get(self._lang_sel, {})  # Actualiza el diccionario
    def __init__(self, args=None):
        super().__init__()
        self._lang_sel = lang_init
        self._dict_sel = _dict[self._lang_sel]

    # Uso de la clase
        f_reg = QFontDatabase.addApplicationFont(":/MononokiNerdFontMono-Regular.ttf")
        self.families = QFontDatabase.applicationFontFamilies(f_reg)
        self.fuente = QFont(self.families[0], 10)
        self.setMinimumSize(600, 400)
        self.setMaximumSize(19200, 10800)
        flags = self.windowFlags()
        flags |= Qt.WindowType.CustomizeWindowHint
        flags |= Qt.WindowType.WindowTitleHint
        flags |= Qt.WindowType.WindowSystemMenuHint
        flags |= Qt.WindowType.WindowCloseButtonHint
        flags |= Qt.WindowType.WindowMaximizeButtonHint
        self.setWindowFlags(flags)
        self.setWindowTitle(self._dict_sel["Title"])
        app_icon = QIcon()
        app_icon.addFile(":IMG/icon32.png", QSize(50, 50))
        self.setWindowIcon(app_icon)
        self.initUI()
        if args is not None and len(args) > 1:  # Acción
            args = args[1:]
            try:
                self.dialogo_abrir(args[0])
                print(f'{self._dict_sel["open"]} {args[0]}')
            except FileNotFoundError:
                QMessageBox.warning(
                    self, f'{self._dict_sel["Warn"]}', f'{self._dict_sel["notExist"]} {args[0]}.'
                )
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Ocurrió un error: {str(e)}.")
            if len(args) > 1:
                if args[1] == "-cld":
                    self.borrar_cargar()
                    print("Código Ensamblado y Cargado en una memoria vaciada.")
                    self.args_run(args)
                elif args[1] == "-ld":
                    self.cargar()
                    print("Código Ensamblado y Cargado.")
                    self.args_run(args)
                else:
                    QMessageBox.warning(
                        self,
                        "Advertencia",
                        "El segundo argumento debe ser '-ld' o 'cld' para compilar el archivo.",
                    )
                    print(
                        "El segundo argumento debe ser '-ld' o 'cld' para compilar el archivo."
                    )

    def args_run(self, args):
        if len(args) <= 2:
            return
        if args[2] == "-r":
            self.ejecutar()
            print("Código ensamblado ejecutado.")
        elif args[2] == "-st":
            if len(args) > 3 and not (args[3].isdigit() and int(args[3]) != 0):
                QMessageBox.warning(
                    self,
                    "Advertencia",
                    f"Argumento inválido '{args[3]}' no es un número entero positivo.",
                )
                print(
                    f"Argumento inválido '{args[3]}' no es un número entero positivo."
                )
                return
            qty = 1 if len(args) == 3 else int(args[3])
            for _ in range(qty):
                self.ejecutar_instruccion()
                print(f"{_+1} paso(s) ejecutado(s).")
        else:
            QMessageBox.warning(
                self, "Advertencia", f"Argumento inválido '{args[2]}', en posición {3}"
            )
            print(f"Argumento inválido '{args[2]}', en posición {3}")
            return

    def resizeEvent(self, event: "QResizeEvent"):  # type: ignore
        self.fuente = QFont(
            self.families[0], min(max(event.size().height() // 80, 10), 14)
        )
        self.fuente_mid = QFont(
            self.families[0], min(max(event.size().height() // 85, 7), 12)
        )
        self.fuente_min = QFont(
            self.families[0], min(max(event.size().height() // 100, 6), 10)
        )
        for _, i in self.mem.items():
            i.table.setFont(self.fuente_mid)
            i.table.horizontalHeader().setFont(self.fuente)
            i.table.verticalHeader().setFont(self.fuente)

        ed_font = self.fuente
        ed_font.setLetterSpacing(QFont.SpacingType.PercentageSpacing, 110)
        self.editor_codigo.editor.lineNumberArea.setFont(self.fuente_mid)
        self.editor_codigo.editor.setFont(ed_font)
        fontMetrics = QFontMetricsF(ed_font)
        spaceWidth = fontMetrics.horizontalAdvance(" ")
        self.editor_codigo.editor.setTabStopDistance(spaceWidth * 4)
        lcd_font = QFont(
            self.families[0], min(max(event.size().height() // 45, 15), 20)
        )
        lcd_font.setLetterSpacing(QFont.SpacingType.PercentageSpacing, 200)
        self.display.display16x2.setFont(lcd_font)
        self.display.display16x2.setMinimumWidth(int(QFontMetricsF(lcd_font).horizontalAdvance(" "))*(config.cols_LCD+1))
        self.display.display16x2.setMaximumWidth(int(QFontMetricsF(lcd_font).horizontalAdvance(" "))*(config.cols_LCD+6))
        
        #spaceWidth = QFontMetricsF(lcd_font).horizontalAdvance(" ")
        #self.display.display16x2.setFixedWidth(int(spaceWidth*16)+30)

        for child in self.registros.findChildren(QWidget):
            child.setFont(self.fuente_mid)
        for child in self.menuBar().findChildren(QWidget):
            child.setFont(self.fuente_mid)
        self.toolbar.setFont(self.fuente_min)
        self.lst.table.setFont(self.fuente_mid)
        self.lst.table.setWordWrap(False)
        super().resizeEvent(event)


        #setCentralWidget
    # region Drag Drop
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith(
                    ".asm"
                ) or url.toLocalFile().lower().endswith(".txt"):
                    event.accept()
                    return
        event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if file_path.endswith(".asm") or url.toLocalFile().lower().endswith(
                    ".txt"
                ):
                    self.nombre_archivo = file_path
                    self.open_proc()
                    return

    # endregion
    def initUI(self):
        self.mode = "start"  # "edit" "run" "loaded"
        self.misc = []
        self.detected_past = None
        self.setFont(self.fuente)

        to_en = 24 if config.composition == 1 else 23
        to_dis = 23 if config.composition == 1 else 24
        to_en2 = 22 if config.lang_init == "esp" else 21
        to_dis2 = 21 if config.lang_init == "esp" else 22

        self.direccion_inicio = "0000"
        self.create_elements()
    
        self.bpoints = self.editor_codigo.editor.breakline
        self.editor_codigo.editor.textChanged.connect(self.texto_modificado)
        self.state_lib = False
        self.monitor.setText(self._dict_sel["txt_mon"])
        self.monitor.setMaximumHeight(100)
        self.monitor.setReadOnly(True)
        self.monitor.setFont(self.fuente)
        self.monitor.setStyleSheet(config.estilo["scrolled_monitor"])

        self.main_tab = QTabWidget(self, tabPosition=QTabWidget.TabPosition.West, movable=True)
        self.main_tab.addTab(self.editor_codigo, self._dict_sel["editor_tab"])
        self.main_tab.addTab(self.lst, self._dict_sel["lst_tab"])
        self.main_tab.addTab(self.display, "Display")

        self.bloque_regSS = QHBoxLayout()
        self.bloque_regSS.addWidget(self.mem["s"], 3)
        self.bloque_regSS.addWidget(self.registros, 1)
        self.bloque_regSS.addWidget(self.portA, 1)

        self.layout_grid = QGridLayout()
        if config.composition == 0:
            self.comp_0()
        else:
            self.comp_1()

        area_trabajo = QWidget()
        styles = config2.styles_fun()
        area_trabajo.setStyleSheet(styles["work_space"])
        area_trabajo.setLayout(self.layout_grid)
        self._ds = {"s": None, "c": None, "d": None}
        self._size = {"s": None, "c": None, "d": None}
        self.Reg_monitor()
        self.setCentralWidget(area_trabajo)
        self.barra_estado = self.statusBar()
        self.setAcceptDrops(True)

        # region     Menus
        # Barra de menús ---------------------------------------------------------------

        self.toolbar = QToolBar("Main Toolbar")
        self.toolbar.setIconSize(QSize(16, 16))
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        spacer2 = QWidget()
        spacer2.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )
        im_tools = [
            "open",
            "save",
            "ld",
            "clr_ld",
            "back",
            "step",
            "next",
            "run",
            "power",
        ]
        txt = self._dict_sel["toolb_txt"]
        self.tools = [QAction()] * len(im_tools)
        fcns = [
            lambda: self.dialogo_abrir(True),
            self.save_fcn,
            self.cargar,
            self.borrar_cargar,
            self.save_fcn,
            self.ejecutar_instruccion,
            self.run_for_bpoint,
            self.ejecutar,
            self.close,
        ]
        for k, _ in enumerate(self.tools):
            self.tools[k] = QAction(QIcon(f":IMG/{im_tools[k]}.png"), txt[k], self)
            self.tools[k].triggered.connect(fcns[k])
            self.toolbar.addAction(self.tools[k])
            if k == 3:
                self.toolbar.addWidget(spacer)
            elif k == 7:
                self.toolbar.addWidget(spacer2)
        self.addToolBar(self.toolbar)
        self.toolbar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)

        # region Menú -----------------------------------------------------------------
        self.nombre_archivo = False
        menu_bar = self.menuBar()
        self.menu_dict = {
            self._dict_sel["menu_txt"]["fl"]["t"]: [
                [self._dict_sel["menu_txt"]["fl"]["open"], "Ctrl+A", lambda: self.dialogo_abrir(True)],
                [self._dict_sel["menu_txt"]["fl"]["save"], "Ctrl+G", lambda: self.save_fcn("save")],
                [self._dict_sel["menu_txt"]["fl"]["save_as"], "Ctrl+Shift+G", lambda: self.save_fcn("como")],
                ["separator"],
                [self._dict_sel["menu_txt"]["fl"]["exit"], "", self.close],
            ],
            self._dict_sel["menu_txt"]["ed"]["t"]: [
                [self._dict_sel["menu_txt"]["ed"]["undo"], "Ctrl+Z", lambda: self.editor_codigo.editor.undo()],
                [self._dict_sel["menu_txt"]["ed"]["redo"], "Ctrl+Y", lambda: self.editor_codigo.editor.redo()],
                ["separator"],
                [self._dict_sel["menu_txt"]["ed"]["cut"], "", lambda: self.focusWidget().cut()], #
                [self._dict_sel["menu_txt"]["ed"]["cpy"], "", lambda: self.focusWidget().copy()], #
                [self._dict_sel["menu_txt"]["ed"]["pst"], "", lambda: self.focusWidget().paste()], #
                ["separator"],
                [self._dict_sel["menu_txt"]["ed"]["add_ind"], "Ctrl+Tab", self.agregar_sangria],
                [self._dict_sel["menu_txt"]["ed"]["del_ind"], "Shift+Tab", self.quitar_sangria],
                ["separator"],
                [self._dict_sel["menu_txt"]["ed"]["comm"], "Ctrl+B", self.comentar],
                [self._dict_sel["menu_txt"]["ed"]["uncomm"], "Ctrl+N", self.descomentar],
            ],
            self._dict_sel["menu_txt"]["ex"]["t"]: [
                [self._dict_sel["menu_txt"]["ex"]["asm_clr_ld"], "Ctrl+U", self.borrar_cargar],
                [self._dict_sel["menu_txt"]["ex"]["asm_ld"], "Ctrl+J", self.cargar],
                [self._dict_sel["menu_txt"]["ex"]["exec"], "Ctrl+K", self.ejecutar],
                [self._dict_sel["menu_txt"]["ex"]["stp"], "Ctrl+L", self.ejecutar_instruccion],
                [self._dict_sel["menu_txt"]["ex"]["bkp"], "", self.run_for_bpoint],
                ["separator"],
            ],
            self._dict_sel["menu_txt"]["mem"]["t"]: [
                [self._dict_sel["menu_txt"]["mem"]["ld"], "", self.open_file],
                [self._dict_sel["menu_txt"]["mem"]["dmp"], "", self.save_csv],
            ],
            self._dict_sel["menu_txt"]["port"]["t"]: [
                [self._dict_sel["menu_txt"]["port"]["show"], "", lambda: self.toggle()],
            ],
            self._dict_sel["menu_txt"]["lang"]["t"]: [
                [self._dict_sel["menu_txt"]["lang"]["esp"], "", lambda: self.chng_lang("esp")],
                [self._dict_sel["menu_txt"]["lang"]["eng"], "", lambda: self.chng_lang("eng")],
            ],
            self._dict_sel["menu_txt"]["dsg"]["t"]: [
                [self._dict_sel["menu_txt"]["dsg"]["new"], "", lambda: self.chng_dsg("new")],
                [self._dict_sel["menu_txt"]["dsg"]["old"], "", lambda: self.chng_dsg("old")],
            ]
        }
        self.menu = {}
        self.menu_elems = [QAction()]*25
        cnt = 0
        for k, (key, v) in enumerate(self.menu_dict.items()):
            self.menu[k] = menu_bar.addMenu(key)
            for i in v:
                if i == ["separator"]:
                    self.menu[k].addSeparator()
                else:
                    self.menu_elems[cnt] = QAction(i[0], self)
                    if i[1] != "":
                        self.menu_elems[cnt].setShortcut(i[1])
                    self.menu_elems[cnt].triggered.connect(i[2])
                    self.menu[k].addAction(self.menu_elems[cnt])
                    cnt += 1
        self.menu_elems[to_dis].setEnabled(False)
        self.menu_elems[to_en].setEnabled(True)
        self.menu_elems[to_dis2].setEnabled(False)
        self.menu_elems[to_en2].setEnabled(True)
        # endregion

        # region ToolBar
        self.state_def(st_comp=True, st_cnt=False, st_cnt_bkp=False, st_edit=True)

    def create_elements(self):
        self.chkbx = {
            "s": QCheckBox(text=" "),
            "c": QCheckBox(text=" "),
            "d": QCheckBox(text=" "),
        }
        self.datalst = [["","",""]]
        self.lst = lst_table()
        self.editor_codigo = EditorCodigo()
        self.registros = EditorRegistros()
        self.display = LCD()
        self.mem = {
            "s": memory("stack", self),
            "c": memory("code", self),
            "d": memory("data", self),
        }
        self.portA = IOPortA()
        self.monitor = QTextEdit(self)

    def comp_1(self):
        col1 = 3
        col2 = 2
        row1 = 2
        row2 = 3
        row3 = 1
        self.layout_grid.addWidget(self.main_tab,0,0,row1+row2+row3,col1)
        self.layout_grid.addWidget(self.registros,row1+row2+row3,0,row1,col1)
        self.splitter_create(row1, row2, row3)
        self.layout_grid.addWidget(self.splitter,0,col1,row1+row2+row3,col2)
        self.layout_grid.addWidget(self.portA,0,col1+2,row1+row2+row3,1)
        self.layout_grid.addWidget(self.monitor,row1+row2+row3,col1,row1,col2+1)
        self.layout_grid.setColumnStretch(5,0)
        self.layout_grid.setRowMinimumHeight(5,0)
        self.layout_grid.setVerticalSpacing(0)
        for i in range(self.layout_grid.columnCount()):
            self.layout_grid.setColumnStretch(i, 1)
        for i in range(self.layout_grid.rowCount()):
            self.layout_grid.setRowStretch(i, 1)
        self.layout_grid.setColumnStretch(5,0)
        self.layout_grid.setRowMinimumHeight(5,0)

    def splitter_create(self, row1, row2, row3):
        self.splitter = QSplitter(Qt.Orientation.Vertical)
        self.splitter.addWidget(self.mem["d"]) # 0,col1,row1,col2
        self.splitter.addWidget(self.mem["c"]) # row1,col1,row2,col2)
        self.splitter.addWidget(self.mem["s"]) # row1+row2,col1,row3,col2
        self.splitter.setStretchFactor(0, row1)
        self.splitter.setStretchFactor(1, row2)
        self.splitter.setStretchFactor(2, row3)
        self.splitter.setStyleSheet("""
            QSplitter::handle {
                background: qlineargradient(
                    x1: 0, x2: 1,
                    stop: 0 #333333, stop: 0.5 #202020, stop: 1 #333333
                );
                height: 3.5px;
                margin-left: 30px;  /* Márgen izquierdo */
                margin-right: 30px; /* Márgen derecho */
            }
        """)
        self.splitter.setOpaqueResize(True)

    def comp_0(self):
        col1 = 4
        self.layout_grid.addWidget(self.main_tab,0,0,4,col1)
        self.layout_grid.addWidget(self.mem["c"],4,0,2,col1)
        self.layout_grid.addWidget(self.mem["d"],6,0,2,col1)
        self.layout_grid.addLayout(self.bloque_regSS,0,col1,6,2)
        self.layout_grid.addWidget(self.monitor,6,col1,2,2)
        for i in range(self.layout_grid.columnCount()):
            self.layout_grid.setColumnStretch(i, 1)
        for i in range(self.layout_grid.rowCount()):
            self.layout_grid.setRowStretch(i, 1)

    # endregion
    # region FUN Extras

    def extraer_valores(self):  # Detecta las direcciones y directivas ORG
        regex_org = r"\.org\s+0x?([0-9A-Fa-f]*|0)"
        regex_seg = r"\.(dseg|cseg)"
        texto = self.editor_codigo.editor.toPlainText()
        rowed = enumerate(texto.splitlines())
        for i, linea in rowed:
            if match_org := re.search(regex_org, linea):
                valor_hex = int(match_org[1].zfill(4), 16) // 16
                if match_seg := re.search(
                    regex_seg, texto.splitlines()[i + 1]
                ) or re.search(regex_seg, texto.splitlines()[i - 1]):
                    secc = match_seg.group(1)[0]
                else:
                    secc = "s"
                self._ds[secc] = valor_hex

    def ds_op(self):
        _ds_ordered = dict(
            sorted(
                self._ds.items(),
                key=lambda item: (item[1] is None, item[1]),
                reverse=True,
            )
        )
        _max = 4096
        for k, v in _ds_ordered.items():
            if v is None:
                self._size[k] = 0
            else:
                self._size[k] = _max - v
                _max = v
        for (k, v), (_, u) in zip(self.mem.items(), self._size.items()):
            if k == "s" and config.composition == 0:
                v.table.setColumnCount(2)
                if u == 0 and k == "s":
                    v.table.setColumnCount(0)
                    if hasattr(self, "splitter"):
                        self.splitter.setStretchFactor(2,0)
                else:
                    v.table.setHorizontalHeaderLabels(
                        [
                            format(i * 16, "X").zfill(4)
                            for i in range(self._ds[k], self._ds[k] + u)
                        ]
                    )
            else:
                v.table.setRowCount(u)
                if self._ds[k] is not None:
                    v.table.setVerticalHeaderLabels(
                        [
                            format(i * 16, "X").zfill(4)
                            for i in range(self._ds[k], self._ds[k] + u)
                        ]
                    )

    def regen_all(self):
        for v in self.mem.values():
            for i in range(v.table.rowCount()):
                v.table.setRowHeight(i, 8)
                for j in range(v.table.columnCount()):
                    if v.table.item(i, j) is None:
                        v.table.setItem(i, j, QTableWidgetItem("00"))
                        v.table.item(i, j).setTextAlignment(
                            Qt.AlignmentFlag.AlignCenter
                        )

    def update_segments(self, mp_el: dict):
        _ds_ordered = dict(
            sorted(
                self._ds.items(),
                key=lambda item: (item[1] is None, item[1]),
                reverse=True,
            )
        )
        _ds_ordered = [[k, v * 16] for k, v in _ds_ordered.items() if v is not None]
        for ix, val in mp_el.items():
            for qty in _ds_ordered:
                if ix >= qty[1]:
                    i, j = (
                        (ix % 16, ix // 16 - qty[1] // 16)
                        if qty[0] == "s" and config.composition == 0
                        else (ix // 16 - qty[1] // 16, ix % 16)
                    )
                    try:
                        if self.mem[qty[0]].table.item(i, j).text() != val:
                            self.mem[qty[0]].table.item(i, j).setText(val)
                            self.mem[qty[0]].table.item(i, j).setBackground(
                                QColor(255, 75, 75, 90)
                            )
                        else:
                            self.mem[qty[0]].table.item(i, j).setBackground(
                                QColor(20, 20, 20)
                            )
                            if qty[0] == "c":
                                self.mem[qty[0]].table.item(i, j).setForeground(
                                    QColor(120, 150, 175)
                                )
                        break
                    except:
                        print(i, j, qty[0])

    def F_monitor(self):
        bool_flags = [x.text() == "1" for x in self.registros.edit_banderas]
        for i, flag in enumerate(bool_flags):
            if flag:
                self.registros.edit_banderas[i].setStyleSheet(
                    "border: 2px solid rgb(255,60,140);"
                )
            else:
                self.registros.edit_banderas[i].setStyleSheet(
                    "border: 2px solid rgb(0,60,140);"
                )

    def Reg_monitor(self):
        forks = [config.IX, config.IY, config.PP]
        fork_abc = [config.AcA, config.AcB, config.AcC]
        for i, x in enumerate(self.registros.edit_punteros):
            if int(x.text(), 16) == forks[i]:
                x.setStyleSheet("border: 2px solid rgb(255,255,255);")
            else:
                x.setStyleSheet("border: 2px solid rgb(255,60,140);")

        if int(self.registros.edit_PIns.text(), 16) == config.PIns:
            self.registros.edit_PIns.setStyleSheet(
                "border: 2px solid rgb(255,255,255);"
            )
        else:
            self.registros.edit_PIns.setStyleSheet("border: 2px solid rgb(255,60,140);")

        for i, x in enumerate(self.registros.edit_acumuladores):
            if hex_a_op(x.text()) == fork_abc[i]:
                x.setStyleSheet("border: 2px solid rgb(255,255,255);")
            else:
                x.setStyleSheet("border: 2px solid rgb(255,60,140);")

    def color_Regs(self):
        self.Reg_monitor()
        fork = [config.C, config.V, config.H, config.N, config.Z, config.P]
        for i, x in enumerate(self.registros.edit_banderas):
            if fork[i] == 1:
                x.setStyleSheet("border: 2px solid rgb(255,60,140);")
            else:
                x.setStyleSheet("border: 2px solid rgb(0,60,140);")

    # endregion

    # region FUNCIONES DEL MENÚ ARCHIVO ---------------------------------------------------

    def dialogo_abrir(self, cust_name=True):
        nombre_archivo = (
            QFileDialog.getOpenFileName(self, self._dict_sel["open_wndw"])[0]
            if cust_name
            else cust_name
        )
        if nombre_archivo:
            self.nombre_archivo = nombre_archivo
            self.open_proc()
            self.setWindowTitle(f"{self._dict_sel['Title']} - {self.nombre_archivo}")

    def open_proc(self):
        try:
            f = open(self.nombre_archivo, "r", encoding="utf-8")
            with f:
                datos_archivo = f.read()
                self.editor_codigo.editor.setPlainText(datos_archivo)
        except UnicodeDecodeError as e:
            self.decode_latin(f, e)

    def decode_latin(self, f, e):
        print(f"Warn: Latin-1. {e}")
        with open(self.nombre_archivo, "r", encoding="latin-1") as archivo_original:
            contenido = archivo_original.read()
        with open(f'{self.nombre_archivo[:-4]}_UTF8{self.nombre_archivo[-4:]}', "w", encoding="utf-8") as archivo_utf8:
            archivo_utf8.write(contenido)
            self.nombre_archivo = f'{self.nombre_archivo[:-4]}_UTF8{self.nombre_archivo[-4:]}'
            print(f'{self._dict_sel["utf8_stt"]}')
            f.close()
        f = open(self.nombre_archivo, "r", encoding="utf-8")
        with f:
            datos_archivo = f.read()
            self.editor_codigo.editor.setPlainText(datos_archivo)
        self.barra_estado.showMessage(self._dict_sel["utf8_stt"])
        

    def save_fcn(self, save_type: str):
        cursor = self.editor_codigo.editor.textCursor()
        cursor.movePosition(cursor.MoveOperation.End, cursor.MoveMode.MoveAnchor)
        cursor.movePosition(cursor.MoveOperation.Left, cursor.MoveMode.KeepAnchor)
        if cursor.selectedText() in letras_numeros:
            cursor.movePosition(cursor.MoveOperation.End)
            linea_nueva = "\n"
            cursor.insertText(linea_nueva)
        if save_type == "como":
            nombre_archivo = QFileDialog.getSaveFileName(
                self,
                self._dict_sel["save_wndw"],
                "",
                self._dict_sel["asm,all"],
            )
            if not nombre_archivo[0]:
                return
            self.nombre_archivo = nombre_archivo[0]
            self.setWindowTitle(f"{self._dict_sel['Title']} - {nombre_archivo[0]}")
        if self.nombre_archivo:
            nombre_archivo = str(self.nombre_archivo)
            with open(nombre_archivo, "w", encoding="utf-8") as f:
                datos_archivo = self.editor_codigo.editor.toPlainText()
                f.write(datos_archivo)
            self.barra_estado.showMessage(self._dict_sel["saved_stt"])
        else:
            self.save_fcn("como")

    # endregion

    # region FUNCIONES DEL MENÚ EDITAR-----------------------------------------------------

    def agregar_sangria(self):
        tab = "\t"
        cursor = self.editor_codigo.editor.textCursor()
        inicio_seleccion = cursor.selectionStart()
        final_seleccion = cursor.selectionEnd()

        cursor.setPosition(final_seleccion)
        linea_final = cursor.blockNumber()

        cursor.setPosition(inicio_seleccion)
        linea_inicial = cursor.blockNumber()

        for _ in range(linea_inicial, linea_final + 1):
            self.movpos(cursor, tab)

    def quitar_sangria(self):
        tab = "\t"
        cursor = self.editor_codigo.editor.textCursor()
        inicio_seleccion = cursor.selectionStart()
        final_seleccion = cursor.selectionEnd()

        cursor.setPosition(final_seleccion)
        linea_final = cursor.blockNumber()

        cursor.setPosition(inicio_seleccion)
        linea_inicial = cursor.blockNumber()

        for _ in range(linea_inicial, linea_final + 1):
            self.set_curs(cursor, tab)

    def comentar(self):
        punto_coma = ";"
        cursor = self.editor_codigo.editor.textCursor()
        inicio_seleccion = cursor.selectionStart()
        final_seleccion = cursor.selectionEnd()

        cursor.setPosition(final_seleccion)
        linea_final = cursor.blockNumber()

        cursor.setPosition(inicio_seleccion)
        linea_inicial = cursor.blockNumber()

        for _ in range(linea_inicial, linea_final + 1):
            self.movpos(cursor, punto_coma)

    def movpos(self, cursor: QTextCursor, arg1):
        cursor.movePosition(cursor.movePosition.StartOfLine)
        cursor.insertText(arg1)
        cursor.movePosition(cursor.movePosition.Down)

    def descomentar(self):
        punto_coma = ";"
        cursor = self.editor_codigo.editor.textCursor()
        inicio_seleccion = cursor.selectionStart()
        final_seleccion = cursor.selectionEnd()

        cursor.setPosition(final_seleccion)
        linea_final = cursor.blockNumber()

        cursor.setPosition(inicio_seleccion)
        linea_inicial = cursor.blockNumber()

        for _ in range(linea_inicial, linea_final + 1):
            self.set_curs(cursor, punto_coma)

    def set_curs(self, cursor: QTextCursor, arg1):
        cursor.movePosition(cursor.movePosition.StartOfLine, cursor.MoveMode.MoveAnchor)
        cursor.movePosition(
            cursor.movePosition.NextCharacter, cursor.MoveMode.KeepAnchor
        )
        if cursor.selectedText() == arg1:
            cursor.deleteChar()
        cursor.movePosition(cursor.movePosition.Down)

    # endregion
    # region FUNCIONES DEL MENÚ EJECUTAR --------------------------------------------------

    def set_Pins(self):
        print(self._ds)
        self.registros.edit_PIns.setText(format(self._ds["c"] * 16, "X").zfill(4))
        config.PIns = int(self.registros.edit_PIns.text(), 16)
        config2.cs_initial = int(self.registros.edit_PIns.text(), 16)

    def borrar_cargar(self):
        self.detected_past = None
        if self.nombre_archivo == False:
            self.save_fcn("como")
        else:
            self.save_fcn("save")
        if self.nombre_archivo:
            cod, err, msj, mp, ls, ts, libs = self.open_toload()
            self.mp = mp.copy()
            self.monitor.setText(msj)
            if err == 0:
                for i in config.m_prog:
                    config.m_prog.update({i: "00"})
                for tab in self.mem.values():
                    tab.reset()
                self.portA.reset()
                self.load(self.nombre_archivo, cod, ls, ts, libs)
                self.state_def(True, True, True, True)

    def open_toload(self):
        with open(self.nombre_archivo, encoding="utf-8") as archivo:
            programa = archivo.readlines()
            cod = list(programa)
        err, msj, mp, ls, ts, libs, self._ds = verificacion_codigo(
                programa, self.nombre_archivo
            )
        
        return cod,err,msj,mp,ls,ts,libs

    def state_def(
        self, st_comp: bool, st_cnt: bool, st_cnt_bkp: bool, st_edit: bool, mems=True
    ):
        self.change_menu(
            13, st_comp, 12, 2
        )
        self.toolbar.actions()[3].setEnabled(st_comp)

        self.change_menu(14, st_cnt, 15, 6)
        self.toolbar.actions()[8].setEnabled(st_cnt)

        self.change_menu(
            16, st_cnt_bkp, 16, 7
        )
        self.editor_codigo.editor.setReadOnly(not st_edit)
        for i in self.mem.values():
            i.table.setEnabled(mems)

    def change_menu(self, arg0, arg1, arg2, arg3):
        self.menu_elems[arg0].setEnabled(arg1)
        self.menu_elems[arg2].setEnabled(arg1)
        self.toolbar.actions()[arg3].setEnabled(arg1)

    def cargar(self):
        if self.nombre_archivo == False:
            self.save_fcn("como")
        else:
            self.save_fcn("save")
        if self.nombre_archivo:
            cod, err, msj, mp, ls, ts, libs = self.open_toload()
            self.mp = mp.copy()
            self.monitor.setText(msj)
            if err == 0:
                self.load(self.nombre_archivo, cod, ls, ts, libs)
                self.state_def(True, True, True, True)

    def load(self, nombre_archivo, cod, ls, ts, libs):
        self.datalst, _ = crear_archivo_listado(nombre_archivo, cod, ls, ts, libs)
        self.rows, self.mem_place = [x[0] for x in self.datalst], [
            x[1] for x in self.datalst
        ]
        self.lst.update(self.datalst)
        config.m_prog.update(self.mp)
        #self.extraer_valores()
        self.ds_op()
        self.regen_all()
        self.update_segments(self.mp)
        self.set_Pins()
        self.post = int(self.registros.edit_PIns.text(), 16) - self._ds["c"] * 16
        self.mode = "coding"
        self.draw_ip()

    def ejecutar(self):
        isd = 0
        while config.PIns != "FIN":
            self.registros.edit_PIns.setText(dec_a_hex4(config.PIns))
            isd += 1
            ciclo_instruccion()
        self.registros.actualizar_registros()
        self.portA.update()
        self.color_Regs()
        self.update_segments(config.m_prog)
        self.post = int(self.registros.edit_PIns.text(), 16) - self._ds["c"] * 16
        if config.PIns == "FIN":
            self.barra_estado.showMessage(self._dict_sel["HLT_stt"])
        self.draw_ip()

    def run_for_bpoint(self):
        bp_dict = dict(zip(self.rows, self.mem_place))
        to_break = {key: bp_dict[str(key)] for key in self.bpoints}
        ktmp, vtmp = list(to_break.keys()), list(to_break.values())
        while config.PIns != "FIN":
            ciclo_instruccion()
            self.registros.actualizar_registros()
            self.portA.update()
            if config.PIns != "FIN":
                pre_ins = f"{int(config.PIns):04X}"
            else:
                self.barra_estado.showMessage(self._dict_sel["HLT_stt"])
            if pre_ins in vtmp:
                msg_bk = (
                    f"{self._dict_sel['brk_stt']} {str(ktmp[vtmp.index(pre_ins)])})"
                )
                self.barra_estado.showMessage(msg_bk)
                break
        self.color_Regs()
        self.update_segments(config.m_prog)
        self.post = int(self.registros.edit_PIns.text(), 16) - self._ds["c"] * 16
        self.draw_ip()

    def ejecutar_instruccion(self):
        if config.PIns != "FIN":
            ciclo_instruccion()
            self.color_Regs()
            self.registros.actualizar_registros()
            self.portA.update()
            self.post = int(self.registros.edit_PIns.text(), 16) - self._ds["c"] * 16
        else:
            self.barra_estado.showMessage(self._dict_sel["HLT_stt"])
        self.update_segments(config.m_prog)
        self.draw_ip()

    def draw_ip(self):
        try:
            self.mem["c"].table.item(self.post // 16, self.post % 16).setBackground(
                QColor(0, 255, 100)
            )
        except:
            print(self.post // 16, self.post % 16)
        self.mem["c"].table.item(self.post // 16, self.post % 16).setForeground(
            QColor(20, 60, 134)
        )
        if self.mode != "loaded":
            if config.PIns != "FIN":
                detected = max(
                    i
                    for i, v in enumerate(self.mem_place)
                    if v == f"{int(config.PIns):04X}"
                )
            else:
                detected = max(
                    i
                    for i, v in enumerate(self.mem_place)
                    if v == self.registros.edit_PIns.text()
                )
            if self.rows[detected] == "lib" and not self.state_lib:
                self.state_lib = True
                self.detected_2 += 1
            elif self.rows[detected] != "lib":
                self.state_lib = False
                self.detected_2 = int(self.rows[detected]) - 1
            if self.mode != "coding-pend":
                self.editor_codigo.editor.highl_IP(self.detected_2)
            for i in range(self.lst.table.columnCount()):
                if self.detected_past is not None:
                    self.lst.table.item(self.detected_past, i).setBackground(
                        QColor(20, 20, 20)
                    )
                    self.lst.table.item(self.detected_past, i).setForeground(
                        QColor(120, 150, 175)
                    )
                self.lst.table.item(detected, i).setBackground(QColor(0, 255, 100))
                self.lst.table.item(detected, i).setForeground(QColor(20, 60, 134))
            self.detected_past = detected

    # endregion
    # region FUNCIONES DEL MENÚ MEMORIA --------------------------------------------------
    def ex2csv(self, f, sh):
        writer = csv.writer(f)
        for r in sh.rows:
            writer.writerow(
                [
                    cell.value
                    for cell in r
                    if cell.value is not None and cell.value != ""
                ]
            )
        return f.getvalue()

    def csv_gen(self, f, strs):
        writer = csv.writer(f)
        # Guardado de ORG
        org_data = ["ORG leaps SS,CS,DS"]
        org_data.extend(str(i) for i in self._ds.values())
        writer.writerow(org_data)
        for k, x in self.mem.items():
            if not self.chkbx[k].isChecked():
                continue
            writer.writerow([strs[k]])
            for i in range(x.table.rowCount()):
                if (
                    any(
                        x.table.item(i, v).text() != "00"
                        for v in range(x.table.columnCount())
                    )
                    or self.response_clear != QMessageBox.StandardButton.Yes
                ):
                    row_data = [str(i)]
                    row_data.extend(
                        x.table.item(i, j).text() for j in range(x.table.columnCount())
                    )
                    writer.writerow(row_data)

    def save_fun(self):
        strs = {"s": "Stack Segment", "c": "Code Segment", "d": "Data Segment"}
        nombre_archivo, tipo_archivo = QFileDialog.getSaveFileName(
            self, self._dict_sel["save_wndw"], "", self._dict_sel["CSV_XLS"]
        )
        buffer = io.StringIO()
        self.csv_gen(buffer, strs)
        with open(nombre_archivo, "w", newline="", encoding="utf-8") as f:
            if tipo_archivo == self._dict_sel["CSV_file"] or nombre_archivo.endswith(".csv"):
                contenido = buffer.getvalue()
                f.write(contenido)
            elif tipo_archivo == self._dict_sel["XLS_file"] or nombre_archivo.endswith(
                ".xlsx"
            ):
                self.save_to_xlsx(buffer, nombre_archivo)
        buffer.close()
        self.msg.accept()
        QMessageBox.information(
            self, self._dict_sel["dump_mssg_t"], self._dict_sel["dump_mssg"]
        )
        self.barra_estado.showMessage(self._dict_sel["dump_stt"])

    def save_to_xlsx(self, buffer, nombre_archivo):
        wb = Workbook()
        ws = wb.active
        buffer.seek(0)
        reader = csv.reader(buffer)
        for row in reader:
            ws.append(row)
        wb.save(nombre_archivo)

    def dialog_save(self, msg_str: str, row=None):
        self.msg = QDialog(self)
        self.msg.setWindowFlags(
            Qt.WindowType.CustomizeWindowHint
            | Qt.WindowType.Dialog
            | Qt.WindowType.WindowTitleHint
        )
        self.msg.setWindowFlags(
            self.msg.windowFlags() & ~Qt.WindowType.WindowCloseButtonHint
        )
        lbl = QLabel(f"{self._dict_sel['seg_select']} {msg_str}:")
        self.btt_dialog = QPushButton(msg_str.upper())
        self.btt_dialog.setEnabled(False)
        self.chkbx["s"] = QCheckBox(text=self._dict_sel["SS"])
        self.chkbx["c"] = QCheckBox(text=self._dict_sel["CS"])
        self.chkbx["d"] = QCheckBox(text=self._dict_sel["DS"])
        if msg_str.upper() == self._dict_sel["imp"].upper():
            self.msg.setWindowTitle(self._dict_sel["ld"])
            self.btt_dialog.clicked.connect(lambda: self.csv2mem(row))
        else:
            self.msg.setWindowTitle(self._dict_sel["dmp"])
            self.btt_dialog.clicked.connect(lambda: self.save_fun)
        layout = QVBoxLayout()
        layout.addWidget(lbl, stretch=1)
        for x, i in self.chkbx.items():
            i.stateChanged.connect(lambda: self.stt_chk(self.chkbx))
            i.setEnabled(self.state[x])
            i.setChecked(self.state[x])
            layout.addWidget(i, stretch=1)
        layout.addWidget(self.btt_dialog, stretch=1)
        self.msg.setLayout(layout)
        self.msg.exec()

    def stt_chk(self, chkbx):
        self.btt_dialog.setEnabled(
            chkbx["s"].isChecked() or chkbx["d"].isChecked() or chkbx["c"].isChecked()
        )

    def save_csv(self):
        self.state = {"s": True, "c": True, "d": True}
        for k, v in self.mem.items():
            self.state[k] = v.table.rowCount() != 0 and v.table.columnCount() != 0
        self.response_clear = None
        mensaje = QMessageBox()
        mensaje.setIcon(QMessageBox.Icon.Question)  # Icono de pregunta
        mensaje.setText(self._dict_sel["omit_zeros"])
        mensaje.setWindowTitle(self._dict_sel["omit_zeros_t"])
        mensaje.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        mensaje.setDefaultButton(QMessageBox.StandardButton.Yes)
        self.response_clear = mensaje.exec()
        self.dialog_save(self._dict_sel["exp"])

    def recon_seg(self, csv):
        rows = list(csv)
        for u in rows:
            for v in range(max(1, len(u) - 1)):
                if u[v].lower() == "stack segment":
                    self.state["s"] = True
                elif u[v].lower() == "code segment":
                    self.state["c"] = True
                elif u[v].lower() == "data segment":
                    self.state["d"] = True
        return rows

    def csv2mem(self, rows):
        mem = None
        for u in rows:
            for v in range(max(1, len(u) - 1)):
                if u[v].lower() in ["stack segment", "code segment", "data segment"]:
                    mem = (
                        self.mem[u[v].lower()[0]].table
                        if self.chkbx[u[v].lower()[0]].isChecked()
                        else None
                    )
                elif u[0].lower() == "org leaps ss,cs,ds":
                    for k, val in self.state.items():
                        if val:
                            self._ds[k] = int(u[self._ds.index(k) + 1])
                    self.ds_op()
                    self.regen_all()
                    break
                elif mem is not None:
                    mem.item(int(u[0]), v).setText(u[v + 1].zfill(2).upper())
        QMessageBox.information(
            self, self._dict_sel["ld_mssg_t"], self._dict_sel["ld_mssg"]
        )
        self.barra_estado.showMessage(self._dict_sel["ld_stt"])
        self.state_def(True, True, False, True)
        self.mode = "loaded"
        self.msg.accept()

    def open_file(self):
        self.state = {"s": False, "c": False, "d": False}
        nombre_archivo, tipo_archivo = QFileDialog.getOpenFileName(
            self, self._dict_sel["open_wndw"], "", self._dict_sel["csvxls_both"]
        )
        for _, i in self.mem.items():
            i.table.setEnabled(True)
        try:
            if tipo_archivo == self._dict_sel["CSV_file"] or nombre_archivo.endswith(".csv"):
                csv_reader = csv.reader(open(nombre_archivo, "r", encoding="utf-8"))
                rows = self.recon_seg(csv_reader)
                self.dialog_save(self._dict_sel["imp"], rows)
            elif tipo_archivo == self._dict_sel["XLS_file"] or nombre_archivo.endswith(
                ".xlsx"
            ):
                self.open_from_xlsx(nombre_archivo)
            self.set_Pins()
        except PermissionError:
            QMessageBox.critical(
                self,
                "Error",
                self._dict_sel["read_eRR"],
            )
            return

    def open_from_xlsx(self, nombre_archivo):
        wb = load_workbook(nombre_archivo)
        sh = wb.active
        buffer2 = io.StringIO()
        content = self.ex2csv(buffer2, sh)
        open("tmp.$csv", "w", encoding="utf-8").write(content)
        csv_reader = csv.reader(open("tmp.$csv", "r", encoding="utf-8"))
        self.recon_seg(csv_reader)
        self.dialog_save(self._dict_sel["imp"], csv_reader)
        os.remove("tmp.$csv")
        buffer2.close()

    def toggle(self):
        self.portA.setVisible(not self.portA.isVisible())
        self.layout_grid.setColumnStretch(5,1 if self.portA.isVisible() else 0)
        self.layout_grid.setRowMinimumHeight(5,1 if self.portA.isVisible() else 0)
        return

    def chng_lang(self, lang: str):
        past = self.lang_sel
        self.lang_sel = lang
        config.lang = lang
        if self.nombre_archivo:
            self.setWindowTitle(f"{self._dict_sel['Title']} - {self.nombre_archivo}")
        else:
            self.setWindowTitle(self._dict_sel["Title"])
        to_en = 22 if lang == "esp" else 21
        to_dis = 21 if lang == "esp" else 22
        self.menu_elems[to_dis].setEnabled(False)
        self.menu_elems[to_en].setEnabled(True)
        if self.monitor.toPlainText() == _dict[past]["txt_mon"]:
            self.monitor.setText(self._dict_sel["txt_mon"])
        self.main_tab.setTabText(0, self._dict_sel["editor_tab"])
        self.main_tab.setTabText(1, self._dict_sel["lst_tab"])
        txt = self._dict_sel["toolb_txt"]
        for k, _ in enumerate(self.tools):  # Recorremos del 1 al 10
            self.tools[k].setText(txt[k])
        men_txt = [self._dict_sel["menu_txt"]["fl"]["t"],
                self._dict_sel["menu_txt"]["ed"]["t"],
                self._dict_sel["menu_txt"]["ex"]["t"],
                self._dict_sel["menu_txt"]["mem"]["t"],
                self._dict_sel["menu_txt"]["port"]["t"],
                self._dict_sel["menu_txt"]["lang"]["t"],
                self._dict_sel["menu_txt"]["dsg"]["t"]]
        menu_dict_txt = [
                self._dict_sel["menu_txt"]["fl"]["open"],
                self._dict_sel["menu_txt"]["fl"]["save"],
                self._dict_sel["menu_txt"]["fl"]["save_as"],
                self._dict_sel["menu_txt"]["fl"]["exit"],
                self._dict_sel["menu_txt"]["ed"]["undo"],
                self._dict_sel["menu_txt"]["ed"]["redo"],
                self._dict_sel["menu_txt"]["ed"]["cut"],
                self._dict_sel["menu_txt"]["ed"]["cpy"],
                self._dict_sel["menu_txt"]["ed"]["pst"],
                self._dict_sel["menu_txt"]["ed"]["add_ind"],
                self._dict_sel["menu_txt"]["ed"]["del_ind"],
                self._dict_sel["menu_txt"]["ed"]["comm"],
                self._dict_sel["menu_txt"]["ed"]["uncomm"],
                self._dict_sel["menu_txt"]["ex"]["asm_clr_ld"],
                self._dict_sel["menu_txt"]["ex"]["asm_ld"],
                self._dict_sel["menu_txt"]["ex"]["exec"],
                self._dict_sel["menu_txt"]["ex"]["stp"],
                self._dict_sel["menu_txt"]["ex"]["bkp"],
                self._dict_sel["menu_txt"]["mem"]["ld"],
                self._dict_sel["menu_txt"]["mem"]["dmp"],
                self._dict_sel["menu_txt"]["port"]["show"],
                self._dict_sel["menu_txt"]["lang"]["esp"],
                self._dict_sel["menu_txt"]["lang"]["eng"],
                self._dict_sel["menu_txt"]["dsg"]["new"],
                self._dict_sel["menu_txt"]["dsg"]["old"]]
        for k, _ in enumerate(self.menu):
            self.menu[k].setTitle(men_txt[k])
        for k, _ in enumerate(self.menu_elems):
            self.menu_elems[k].setText(menu_dict_txt[k])
        self.registros.upd_lang(self.lang_sel)
        self.lst.upd_lang(self.lang_sel)
        self.portA.upd_lang(self.lang_sel)
        self.mem["s"].upd_lang(self.lang_sel)
        self.mem["c"].upd_lang(self.lang_sel)
        self.mem["d"].upd_lang(self.lang_sel)

    def chng_dsg(self, dsg: str):
        to_en = 24 if dsg == "new" else 23
        to_dis = 23 if dsg == "new" else 24
        self.menu_elems[to_dis].setEnabled(False)
        self.menu_elems[to_en].setEnabled(True)
        while self.layout_grid.count() > 0:
            item = self.layout_grid.takeAt(0)  # Tomar el primer elemento del layout
            if widget := item.widget():
                widget.setParent(None)   # Quitar el widget del layout sin destruirlo
        if dsg == "old":
            self.comp_0()
        else:
            self.comp_1()
        config.composition = 1 if dsg == "new" else 0
        self.registros.redraw()
    
    def _upd_LCD(self, data, pos_ini):
        self.display.update(data, pos_ini)
        


    # FUNCIONES DEL EDITOR DE CÓDIGO -----------------------------------------------

    def texto_modificado(self):
        self.mode = "coding-pend"
        self.barra_estado.showMessage(self._dict_sel["not_saved_stt"])


if __name__ == "__main__":

    app = QApplication(sys.argv)
    ex = ComputadorCompleto(sys.argv)

    ex.show()
    sys.exit(app.exec())
